# -*- coding: utf-8 -*-
"""Motor de presupuesto estilo VEGRA. Tarifa reconstruida de la oferta V-26143."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

AR = "Arial"
EUR = '#,##0.00\\ "€"'; MM = '#,##0'
FACTORS = [("rec", "Conducte recte", 1.00), ("red", "Conducte reducció", 1.03),
           ("c90", "Corba 90°", 1.00), ("c45", "Corba 45°", 0.80),
           ("inj", "Injert / derivació", 0.50), ("des", "Desviament", 1.02)]
FACT_RANGE = "Tarifa!$A$10:$B$15"

def _styles():
    return dict(
        hdr=PatternFill("solid", fgColor="1F3864"), grp=PatternFill("solid", fgColor="D9E1F2"),
        sub=PatternFill("solid", fgColor="EEF2FB"), inp=PatternFill("solid", fgColor="FFF2CC"),
        tot=PatternFill("solid", fgColor="C6E0B4"),
        box=Border(*[Side(style="thin", color="BBBBBB")] * 4),
        blue=Font(name=AR, color="0000FF", size=10), blk=Font(name=AR, color="000000", size=10))

def build_budget(groups, labor, out_path, base=21.28, tapa=10.64, malla=63.83, iva=0.21, materials=None):
    """groups: list of (titulo, [ (codi,descr,w1,h1,w2,h2,uts,unit,preu_manual|None) ])
       labor:  list of (descr, importe)"""
    S = _styles()
    wb = Workbook(); t = wb.active; t.title = "Tarifa"
    t["A1"] = "TARIFA — reconstruïda del pressupost V-26143 (VEGRA 400)"
    t["A1"].font = Font(name=AR, bold=True, size=12)
    rows = [("Base xapa simple galva (€/m²)", base), ("Tapa simple (€/ut)", tapa),
            ("Tapa malla (€/ut)", malla), ("IVA", iva)]
    for i, (k, v) in enumerate(rows, start=3):
        t[f"A{i}"] = k; t[f"A{i}"].font = Font(name=AR, bold=True, size=10)
        t[f"B{i}"] = v; t[f"B{i}"].font = S["blue"]; t[f"B{i}"].fill = S["inp"]; t[f"B{i}"].border = S["box"]
    t["B6"].number_format = "0%"
    t["B3"].comment = Comment("Derivat dels conductes rectes de la V-26143 (€/m² de xapa). Ajustable.", "sistema")
    t["A8"] = "Factor per tipus de peça"; t["A8"].font = Font(name=AR, bold=True, size=10)
    t["A9"], t["B9"], t["C9"] = "codi", "factor", "descripció"
    for c in ("A9", "B9", "C9"): t[c].font = Font(name=AR, bold=True, size=10); t[c].fill = S["sub"]
    for i, (code, desc, f) in enumerate(FACTORS, start=10):
        t[f"A{i}"] = code; t[f"B{i}"] = f; t[f"C{i}"] = desc
        t[f"B{i}"].font = S["blue"]; t[f"B{i}"].fill = S["inp"]
        for col in "ABC": t[f"{col}{i}"].border = S["box"]
    t.column_dimensions["A"].width = 28; t.column_dimensions["B"].width = 10; t.column_dimensions["C"].width = 22

    p = wb.create_sheet("Pressupost")
    p["A1"] = "PRESSUPOST AUTOMÀTIC"; p["A1"].font = Font(name=AR, bold=True, size=13)
    p["A2"] = "Generat des del plano · tarifa a full 'Tarifa' (editable)"
    p["A2"].font = Font(name=AR, italic=True, size=9, color="666666")
    heads = ["Descripció", "Tipus", "W1", "H1", "W2", "H2", "Perím (m)", "UTS", "Unitat", "Preu (€)", "Subtotal (€)"]
    hrow = 4
    for j, h in enumerate(heads, 1):
        c = p.cell(hrow, j, h); c.font = Font(name=AR, bold=True, color="FFFFFF", size=10)
        c.fill = S["hdr"]; c.alignment = Alignment(horizontal="center"); c.border = S["box"]
    r = hrow + 1; subtotal_rows = []
    for titulo, items in groups:
        c = p.cell(r, 1, titulo); c.font = Font(name=AR, bold=True, size=11)
        for col in range(1, 12): p.cell(r, col).fill = S["grp"]; p.cell(r, col).border = S["box"]
        r += 1; first = r
        for (code, descr, w1, h1, w2, h2, uts, unit, manual) in items:
            p.cell(r, 1, descr); p.cell(r, 2, code)
            if code != "esp":
                p.cell(r, 3, w1); p.cell(r, 4, h1); p.cell(r, 5, w2); p.cell(r, 6, h2)
                p.cell(r, 7, f"=((2*(C{r}+D{r}))+(2*(E{r}+F{r})))/2/1000")
            p.cell(r, 8, uts); p.cell(r, 9, unit)
            if manual is not None: p.cell(r, 10, manual)
            elif code == "tapa": p.cell(r, 10, "=Tarifa!$B$4")
            elif code == "tmalla": p.cell(r, 10, "=Tarifa!$B$5")
            else: p.cell(r, 10, f"=G{r}*Tarifa!$B$3*VLOOKUP(B{r},{FACT_RANGE},2,0)")
            p.cell(r, 11, f"=H{r}*J{r}")
            for col in range(1, 12): p.cell(r, col).border = S["box"]; p.cell(r, col).font = S["blk"]
            if manual is not None: p.cell(r, 10).font = S["blue"]; p.cell(r, 10).fill = S["inp"]
            p.cell(r, 7).number_format = "0.000"
            for col in (10, 11): p.cell(r, col).number_format = EUR
            for col in (3, 4, 5, 6): p.cell(r, col).number_format = MM
            r += 1
        p.cell(r, 1, "  Subtotal").font = Font(name=AR, bold=True, italic=True, size=10)
        sc = p.cell(r, 11, f"=SUM(K{first}:K{r-1})"); sc.font = Font(name=AR, bold=True, size=10); sc.number_format = EUR
        for col in range(1, 12): p.cell(r, col).fill = S["sub"]; p.cell(r, col).border = S["box"]
        subtotal_rows.append(r); r += 1
    r += 1
    matrow = r
    p.cell(r, 1, "SUBTOTAL MATERIAL").font = Font(name=AR, bold=True, size=11)
    mc = p.cell(r, 11, "=" + "+".join(f"K{x}" for x in subtotal_rows))
    mc.font = Font(name=AR, bold=True, size=11); mc.number_format = EUR
    for col in range(1, 12): p.cell(r, col).fill = S["tot"]; p.cell(r, col).border = S["box"]
    r += 2
    p.cell(r, 1, "2 · TREBALLS (mà d'obra)").font = Font(name=AR, bold=True, size=11)
    for col in range(1, 12): p.cell(r, col).fill = S["grp"]; p.cell(r, col).border = S["box"]
    r += 1; lfirst = r
    for descr, val in labor:
        p.cell(r, 1, descr); lc = p.cell(r, 11, val); lc.font = S["blue"]; lc.fill = S["inp"]; lc.number_format = EUR
        for col in range(1, 12): p.cell(r, col).border = S["box"]
        r += 1
    p.cell(r, 1, "SUBTOTAL TREBALLS").font = Font(name=AR, bold=True, size=10)
    ls = p.cell(r, 11, f"=SUM(K{lfirst}:K{r-1})" if labor else 0); ls.font = Font(name=AR, bold=True); ls.number_format = EUR
    for col in range(1, 12): p.cell(r, col).fill = S["sub"]; p.cell(r, col).border = S["box"]
    trebrow = r; r += 2
    p.cell(r, 1, "SUBTOTAL BASE").font = Font(name=AR, bold=True, size=11)
    bc = p.cell(r, 11, f"=K{matrow}+K{trebrow}"); bc.font = Font(name=AR, bold=True, size=11); bc.number_format = EUR
    baserow = r
    for col in range(1, 12): p.cell(r, col).border = S["box"]
    r += 1
    p.cell(r, 1, "IVA 21%").font = Font(name=AR, bold=True)
    ic = p.cell(r, 11, f"=K{baserow}*Tarifa!$B$6"); ic.number_format = EUR; ivarow = r; r += 1
    for col in range(1, 12): p.cell(ivarow, col).border = S["box"]
    p.cell(r, 1, "TOTAL OFERTA").font = Font(name=AR, bold=True, size=12)
    fc = p.cell(r, 11, f"=K{baserow}+K{ivarow}"); fc.font = Font(name=AR, bold=True, size=12); fc.number_format = EUR
    for col in range(1, 12): p.cell(r, col).fill = S["tot"]; p.cell(r, col).border = S["box"]
    r += 2
    p.cell(r, 1, "Llegenda: blau/groc = dada d'entrada (peça especial, mà d'obra o tarifa) · negre = calculat per fórmula").font = Font(name=AR, italic=True, size=8, color="666666")
    for j, wd in enumerate([42, 7, 7, 7, 7, 7, 10, 8, 8, 12, 13], 1):
        p.column_dimensions[p.cell(1, j).column_letter].width = wd
    p.freeze_panes = "A5"
    if materials:
        _materials_sheet(wb, materials, S)
    wb.save(out_path)
    return out_path

def _materials_sheet(wb, m, S):
    ms = wb.create_sheet("Materials")
    ms["A1"] = "MATERIALS I ACCESSORIS (estimat del despiece)"
    ms["A1"].font = Font(name=AR, bold=True, size=12)
    def row(r, a, b, c="", note="", inp=False, fmt=None, bold=False):
        ms[f"A{r}"] = a; ms[f"A{r}"].font = Font(name=AR, bold=bold, size=10)
        cell = ms[f"B{r}"]; cell.value = b
        cell.font = S["blue"] if inp else S["blk"]
        if inp: cell.fill = S["inp"]
        if fmt: cell.number_format = fmt
        ms[f"C{r}"] = c; ms[f"C{r}"].font = Font(name=AR, size=9, color="666666")
        if note: ms[f"D{r}"] = note; ms[f"D{r}"].font = Font(name=AR, italic=True, size=8, color="888888")
    row(3, "PARÀMETRES (editables)", "", "", bold=True)
    row(4, "Merma de xapa", 0.12, "%", "retalls", inp=True, fmt="0%")
    row(5, "Pes xapa 0,8 mm", 6.4, "kg/m²", "galvanitzat", inp=True, fmt="0.0")
    row(6, "Suport/abraçadera cada", 3, "m", "cada 3 m", inp=True, fmt="0.0")
    row(7, "Cargol cada", 200, "mm de brida", "metu", inp=True, fmt="0")
    row(9, "QUANTITATS (del despiece)", "", "", bold=True)
    row(10, "Superfície xapa neta", m["area_neta_m2"], "m²", inp=False, fmt="0.0")
    row(11, "Metres de conducte", m["metres"], "m", fmt="0.0")
    row(12, "Nº de peces (trams/accessoris)", m["peces"], "ut", fmt="0")
    row(13, "Perímetre total de brides", m["brides_mm"], "mm", fmt="#,##0")
    row(15, "RESULTATS", "", "", bold=True)
    ms["A16"] = "Superfície amb merma"; ms["A16"].font = Font(name=AR, bold=True, size=10)
    ms["B16"] = "=B10*(1+B4)"; ms["B16"].number_format = "0.0"; ms["C16"] = "m²"
    ms["A17"] = "Pes de xapa"; ms["A17"].font = Font(name=AR, bold=True, size=10)
    ms["B17"] = "=B16*B5"; ms["B17"].number_format = "#,##0"; ms["C17"] = "kg"
    ms["A18"] = "Abraçaderes"; ms["A18"].font = Font(name=AR, bold=True, size=10)
    ms["B18"] = "=CEILING(B11/B6,1)"; ms["B18"].number_format = "0"; ms["C18"] = "ut"
    ms["A19"] = "Metu (perfil)"; ms["A19"].font = Font(name=AR, bold=True, size=10)
    ms["B19"] = "=B13/1000"; ms["B19"].number_format = "0.0"; ms["C19"] = "m lineals"
    ms["D19"] = "segons secció de cada unió"; ms["D19"].font = Font(name=AR, italic=True, size=8, color="888888")
    ms["A20"] = "Cargols (aprox.)"; ms["A20"].font = Font(name=AR, bold=True, size=10)
    ms["B20"] = "=CEILING(B13/B7,1)"; ms["B20"].number_format = "#,##0"; ms["C20"] = "ut"
    ms["A22"] = "Estimacions a partir del despiece i ratis editables. Les peces especials (pantaló, plenum…) porten el seu material a part."
    ms["A22"].font = Font(name=AR, italic=True, size=8, color="888888")
    for col, w in (("A", 30), ("B", 12), ("C", 14), ("D", 16)):
        ms.column_dimensions[col].width = w
    return ms
